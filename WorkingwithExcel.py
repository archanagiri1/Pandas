"""
Working with Excel (Explanations)
=============================================================
"""

import pandas as pd
import numpy as np
from datetime import datetime

print("=" * 80)
print("WORKING WITH EXCEL -EXPLANATIONS")
print("=" * 80)

# ============================================================================
# PART 1: READING EXCEL FILES - BASIC CONCEPTS
# ============================================================================
print("\n" + "=" * 80)
print("PART 1: READING EXCEL FILES")
print("=" * 80)

"""
WHAT IS AN EXCEL FILE?
======================
- Excel files store data in spreadsheets
- Each file can have multiple sheets (like tabs in a browser)
- Excel files have extensions: .xlsx (new) or .xls (old)
- Pandas can read both formats and convert them to DataFrames

WHY READ EXCEL FILES?
=====================
- Most business data comes in Excel format
- Easy to share and view
- Non-technical users can create Excel files
- Common format for reports and data exports
"""

# Create sample Excel file for demonstration
df_sample = pd.DataFrame({
    'employee_id': [1, 2, 3, 4, 5],
    'name': ['Alice', 'Bob', 'Charlie', 'David', 'Eve'],
    'department': ['HR', 'IT', 'Finance', 'IT', 'HR'],
    'salary': [50000, 60000, 75000, 55000, 70000],
    'join_date': pd.date_range('2020-01-01', periods=5, freq='Y')
})

df_sample.to_excel('/tmp/sample.xlsx', index=False, sheet_name='Employees')
print("\nSample Excel file created for demonstration")

# Example 1: Basic read
print("\n" + "-" * 80)
print("EXAMPLE 1: Basic Read from Excel")
print("-" * 80)

"""
EXPLANATION:
- pd.read_excel() is the main function to read Excel files
- You only need to provide the file path
- Pandas automatically reads the first sheet
- Returns a DataFrame (just like a table in Python)

SYNTAX:
df = pd.read_excel('filename.xlsx')

WHAT HAPPENS:
1. Pandas opens the Excel file
2. Reads the first sheet by default
3. First row becomes column names (headers)
4. Remaining rows become data
5. Creates a DataFrame object
"""

df = pd.read_excel('/tmp/sample.xlsx')
print("\nResult:")
print(df)
print(f"\nData type: {type(df)}")
print(f"Shape: {df.shape} (means {df.shape[0]} rows and {df.shape[1]} columns)")

# Example 2: Specify sheet by name
print("\n" + "-" * 80)
print("EXAMPLE 2: Read Specific Sheet by Name")
print("-" * 80)

"""
EXPLANATION:
- Excel files can have multiple sheets (tabs)
- By default, pandas reads the first sheet
- Use sheet_name parameter to specify which sheet to read
- Sheet names are case-sensitive

WHEN TO USE:
- When your Excel file has multiple sheets
- When you need data from a specific sheet
- When the first sheet is not what you need

SYNTAX:
df = pd.read_excel('file.xlsx', sheet_name='SheetName')
"""

df = pd.read_excel('/tmp/sample.xlsx', sheet_name='Employees')
print("\nReading sheet named 'Employees':")
print(df)

# Example 3: Specify sheet by index
print("\n" + "-" * 80)
print("EXAMPLE 3: Read Sheet by Index (Position Number)")
print("-" * 80)

"""
EXPLANATION:
- Sheets are numbered starting from 0
- First sheet = 0, Second sheet = 1, Third sheet = 2, etc.
- This is useful when you don't know the sheet name
- Or when the sheet name might change but position doesn't

INDEX NUMBERING:
Sheet 1 (first sheet)  → index 0
Sheet 2 (second sheet) → index 1
Sheet 3 (third sheet)  → index 2

SYNTAX:
df = pd.read_excel('file.xlsx', sheet_name=0)  # First sheet
df = pd.read_excel('file.xlsx', sheet_name=1)  # Second sheet
"""

df = pd.read_excel('/tmp/sample.xlsx', sheet_name=0)
print("\nReading first sheet (index 0):")
print(df.head(3))

# Example 4: Read all sheets
print("\n" + "-" * 80)
print("EXAMPLE 4: Read All Sheets at Once")
print("-" * 80)

"""
EXPLANATION:
- sheet_name=None means "read all sheets"
- Returns a dictionary, not a DataFrame
- Dictionary keys = sheet names
- Dictionary values = DataFrames

DICTIONARY STRUCTURE:
{
    'Sheet1': DataFrame1,
    'Sheet2': DataFrame2,
    'Sheet3': DataFrame3
}

HOW TO USE:
1. Get all sheets as dictionary
2. Access each sheet by name
3. Loop through all sheets if needed

SYNTAX:
all_sheets = pd.read_excel('file.xlsx', sheet_name=None)
"""

# Create multi-sheet file
with pd.ExcelWriter('/tmp/multi_demo.xlsx', engine='openpyxl') as writer:
    df_sample.to_excel(writer, sheet_name='Employees', index=False)
    df_sample[['name', 'salary']].to_excel(writer, sheet_name='Salaries', index=False)

all_sheets = pd.read_excel('/tmp/multi_demo.xlsx', sheet_name=None)
print(f"\nSheet names found: {list(all_sheets.keys())}")

for sheet_name, df in all_sheets.items():
    print(f"\n{sheet_name} sheet:")
    print(df.head(2))

# Example 5: Read specific columns
print("\n" + "-" * 80)
print("EXAMPLE 5: Read Only Specific Columns")
print("-" * 80)

"""
EXPLANATION:
- Sometimes you don't need all columns
- usecols parameter lets you choose which columns to read
- Saves memory and processing time
- Two ways: by column names or by Excel column letters

WHY USE THIS:
- Large files with many columns
- You only need a few columns for analysis
- Faster reading and less memory usage

METHOD 1 - By Column Names:
df = pd.read_excel('file.xlsx', usecols=['name', 'salary'])

METHOD 2 - By Excel Column Letters:
df = pd.read_excel('file.xlsx', usecols='A:C')  # Columns A, B, C
"""

# Method 1: By column names
df = pd.read_excel('/tmp/sample.xlsx', usecols=['name', 'salary'])
print("\nMethod 1 - Reading 'name' and 'salary' columns:")
print(df)

# Method 2: By Excel columns
df = pd.read_excel('/tmp/sample.xlsx', usecols='A:C')
print("\nMethod 2 - Reading columns A to C:")
print(df)

# Example 6: Read limited rows
print("\n" + "-" * 80)
print("EXAMPLE 6: Read Only First Few Rows")
print("-" * 80)

"""
EXPLANATION:
- nrows parameter limits how many rows to read
- Useful for previewing large files
- Doesn't include the header row in the count

WHEN TO USE:
- Preview large Excel files (check first 10 rows)
- Testing your code before reading entire file
- When you only need recent/top records

EXAMPLE:
nrows=3 means: Read only 3 rows of data (plus header)

SYNTAX:
df = pd.read_excel('file.xlsx', nrows=5)  # Read only 5 rows
"""

df = pd.read_excel('/tmp/sample.xlsx', nrows=3)
print(f"\nReading only first 3 rows:")
print(df)
print(f"Total rows read: {len(df)}")

# Example 7: Skip rows
print("\n" + "-" * 80)
print("EXAMPLE 7: Skip Rows from Top")
print("-" * 80)

"""
EXPLANATION:
- skiprows parameter skips rows from the top of the file
- Useful when Excel has title/description rows before data
- The row after skipped rows becomes the header

COMMON USE CASE:
Excel files often have:
- Row 1: Report Title
- Row 2: Date Created
- Row 3: Empty
- Row 4: Column Headers (this is what you want)
- Row 5+: Actual data

SYNTAX:
df = pd.read_excel('file.xlsx', skiprows=3)  # Skip first 3 rows

WHAT HAPPENS:
If skiprows=2:
- Rows 1-2 are skipped completely
- Row 3 becomes your column headers
- Row 4+ becomes your data
"""

# Create file with metadata
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.append(['Report Title: Employee Data'])
ws.append(['Created: 2024-01-01'])
ws.append([])  # Empty row
ws.append(['employee_id', 'name', 'salary'])
ws.append([1, 'Alice', 50000])
ws.append([2, 'Bob', 60000])
wb.save('/tmp/with_metadata.xlsx')

df = pd.read_excel('/tmp/with_metadata.xlsx', skiprows=3)
print("\nSkipping first 3 rows (title and metadata):")
print(df)

# Example 8: Read without header
print("\n" + "-" * 80)
print("EXAMPLE 8: Read File Without Header Row")
print("-" * 80)

"""
EXPLANATION:
- header=None tells pandas there are no column names
- Pandas will create default column names: 0, 1, 2, 3, etc.
- Use this when your Excel has no header row
- Or when first row is data, not column names

DEFAULT COLUMN NAMES:
Column 1 → 0
Column 2 → 1
Column 3 → 2
etc.

SYNTAX:
df = pd.read_excel('file.xlsx', header=None)

YOU CAN ALSO PROVIDE CUSTOM NAMES:
df = pd.read_excel('file.xlsx', header=None, names=['ID', 'Name', 'Value'])
"""

df = pd.read_excel('/tmp/sample.xlsx', header=None)
print("\nReading without header (first row becomes data):")
print(df)

# Example 9: Specify data types
print("\n" + "-" * 80)
print("EXAMPLE 9: Control Data Types")
print("-" * 80)

"""
EXPLANATION:
- Pandas automatically guesses data types
- Sometimes it guesses wrong
- dtype parameter lets you specify exact types
- Prevents errors and saves memory

WHY SPECIFY TYPES:
1. Prevent wrong interpretations (e.g., ID as number instead of text)
2. Save memory (int32 instead of int64)
3. Ensure consistency
4. Avoid conversion errors later

COMMON DATA TYPES:
'int32'    - Small integers
'int64'    - Large integers
'float64'  - Decimal numbers
'str'      - Text/strings
'category' - Repeated text values (saves memory)

SYNTAX:
df = pd.read_excel('file.xlsx', dtype={'column_name': 'type'})
"""

print("\nWithout specifying dtype:")
df_auto = pd.read_excel('/tmp/sample.xlsx')
print(df_auto.dtypes)

print("\nWith specified dtype (employee_id as string):")
df_typed = pd.read_excel('/tmp/sample.xlsx', dtype={'employee_id': str})
print(df_typed.dtypes)
print("\nNotice employee_id is now 'object' (string) instead of int64")

# Example 10: Parse dates
print("\n" + "-" * 80)
print("EXAMPLE 10: Parse Date Columns")
print("-" * 80)

"""
EXPLANATION:
- Excel stores dates as numbers internally
- Pandas can convert them to proper datetime objects
- parse_dates parameter specifies which columns are dates
- Makes date operations much easier

WITHOUT PARSING:
- Dates read as strings or numbers
- Can't do date arithmetic
- Can't extract year, month, day easily

WITH PARSING:
- Dates are datetime objects
- Can do: date + 7 days, extract month, filter by date range
- Much more powerful for analysis

SYNTAX:
df = pd.read_excel('file.xlsx', parse_dates=['date_column'])
df = pd.read_excel('file.xlsx', parse_dates=['col1', 'col2'])  # Multiple
"""

print("\nWithout parsing dates:")
df_no_parse = pd.read_excel('/tmp/sample.xlsx')
print(f"join_date type: {df_no_parse['join_date'].dtype}")

print("\nWith parsing dates:")
df_parsed = pd.read_excel('/tmp/sample.xlsx', parse_dates=['join_date'])
print(f"join_date type: {df_parsed['join_date'].dtype}")
print("\nNow you can do date operations:")
print(f"First hire: {df_parsed['join_date'].min()}")
print(f"Latest hire: {df_parsed['join_date'].max()}")

# ============================================================================
# PART 2: WRITING TO EXCEL - DETAILED EXPLANATIONS
# ============================================================================
print("\n" + "=" * 80)
print("PART 2: WRITING TO EXCEL")
print("=" * 80)

"""
WHY WRITE TO EXCEL?
===================
- Share data with non-technical users
- Create reports that others can view
- Export analysis results
- Create formatted spreadsheets
- Backup data in universal format

BASIC CONCEPT:
- You have a DataFrame in Python
- Want to save it as an Excel file
- Use to_excel() method
- Creates .xlsx file on your computer
"""

df_write = pd.DataFrame({
    'id': [1, 2, 3],
    'name': ['Alice', 'Bob', 'Charlie'],
    'score': [85.5, 92.3, 78.9]
})

print("\nDataFrame to write:")
print(df_write)

# Example 1: Basic write
print("\n" + "-" * 80)
print("EXAMPLE 1: Basic Write to Excel")
print("-" * 80)

"""
EXPLANATION:
- to_excel() is the main method
- Provide filename with .xlsx extension
- index=False means don't include row numbers
- Creates a new Excel file

WHAT GETS WRITTEN:
- Column names as first row
- Data in subsequent rows
- Default sheet name is 'Sheet1'

SYNTAX:
df.to_excel('output.xlsx', index=False)

WITH INDEX (index=True or default):
   id    name  score
0   1   Alice   85.5
1   2     Bob   92.3
2   3 Charlie   78.9

WITHOUT INDEX (index=False):
id    name  score
 1   Alice   85.5
 2     Bob   92.3
 3 Charlie   78.9
"""

df_write.to_excel('/tmp/output.xlsx', index=False)
print("✓ File created: output.xlsx")
print("  - No row numbers (index=False)")
print("  - Column names in first row")
print("  - Data in rows below")

# Example 2: Custom sheet name
print("\n" + "-" * 80)
print("EXAMPLE 2: Write with Custom Sheet Name")
print("-" * 80)

"""
EXPLANATION:
- By default, sheet is named 'Sheet1'
- sheet_name parameter lets you choose the name
- Makes your Excel file more professional
- Helps organize data when you have multiple sheets

SYNTAX:
df.to_excel('file.xlsx', sheet_name='Students', index=False)

RESULT:
Excel file with one sheet named 'Students' instead of 'Sheet1'
"""

df_write.to_excel('/tmp/output_named.xlsx', sheet_name='Students', index=False)
print("✓ File created with sheet name: 'Students'")
print("  - More descriptive than 'Sheet1'")
print("  - Professional appearance")

# Example 3: Write with index
print("\n" + "-" * 80)
print("EXAMPLE 3: Include Row Numbers (Index)")
print("-" * 80)

"""
EXPLANATION:
- index=True includes row numbers in Excel
- Creates an extra column with numbers 0, 1, 2, etc.
- Useful when row position matters
- Usually better to exclude it (index=False)

WHEN TO INCLUDE INDEX:
- Row number has meaning
- Need to track original position
- Debugging purposes

WHEN TO EXCLUDE INDEX:
- Row numbers are just 0, 1, 2, 3 (not meaningful)
- Sharing with others (cleaner look)
- Most common use case

COMPARISON:

index=True:           index=False:
   id  name             id  name
0   1  Alice            1  Alice
1   2  Bob              2  Bob
2   3  Charlie          3  Charlie
"""

df_write.to_excel('/tmp/with_index.xlsx', index=True)
print("✓ File created WITH index (row numbers shown)")

df_write.to_excel('/tmp/without_index.xlsx', index=False)
print("✓ File created WITHOUT index (cleaner)")

# ============================================================================
# PART 3: MULTIPLE SHEETS - DETAILED EXPLANATION
# ============================================================================
print("\n" + "=" * 80)
print("PART 3: WORKING WITH MULTIPLE SHEETS")
print("=" * 80)

"""
WHAT ARE MULTIPLE SHEETS?
==========================
- Like tabs in a browser
- One Excel file can have many sheets
- Each sheet is a separate table
- All sheets saved in one .xlsx file

REAL-WORLD EXAMPLES:
- Monthly reports: Jan, Feb, Mar sheets in one file
- Department data: HR, IT, Finance sheets
- Different views: Raw Data, Summary, Charts

WHY USE MULTIPLE SHEETS:
- Keep related data in one file
- Organize different categories
- Separate raw data from analysis
- Easy to navigate for users
"""

# Create sample data for multiple sheets
df_sales = pd.DataFrame({
    'product': ['A', 'B', 'C'],
    'sales': [100, 200, 150]
})

df_profit = pd.DataFrame({
    'product': ['A', 'B', 'C'],
    'profit': [20, 40, 30]
})

df_inventory = pd.DataFrame({
    'product': ['A', 'B', 'C'],
    'quantity': [50, 75, 60]
})

print("\nThree DataFrames to write:")
print("\nSales Data:")
print(df_sales)
print("\nProfit Data:")
print(df_profit)
print("\nInventory Data:")
print(df_inventory)

# Example 1: Write multiple sheets
print("\n" + "-" * 80)
print("EXAMPLE 1: Create Excel with Multiple Sheets")
print("-" * 80)

"""
EXPLANATION:
- Use ExcelWriter to write multiple sheets
- Think of it as opening an Excel file for writing
- Add sheets one by one
- Close the file when done

HOW IT WORKS:
1. Create ExcelWriter object (like opening Excel)
2. Write first DataFrame to sheet 'Sales'
3. Write second DataFrame to sheet 'Profit'
4. Write third DataFrame to sheet 'Inventory'
5. Close (file is saved automatically)

SYNTAX:
with pd.ExcelWriter('file.xlsx', engine='openpyxl') as writer:
    df1.to_excel(writer, sheet_name='Sheet1', index=False)
    df2.to_excel(writer, sheet_name='Sheet2', index=False)
    df3.to_excel(writer, sheet_name='Sheet3', index=False)

RESULT:
One Excel file with 3 tabs:
 file.xlsx
   ├── Sales
   ├── Profit
   └── Inventory
"""

with pd.ExcelWriter('/tmp/multi_sheet.xlsx', engine='openpyxl') as writer:
    df_sales.to_excel(writer, sheet_name='Sales', index=False)
    df_profit.to_excel(writer, sheet_name='Profit', index=False)
    df_inventory.to_excel(writer, sheet_name='Inventory', index=False)

print("✓ File created with 3 sheets:")
print("   Sales")
print("   Profit")
print("   Inventory")

# Example 2: Read all sheets back
print("\n" + "-" * 80)
print("EXAMPLE 2: Read All Sheets from Excel")
print("-" * 80)

"""
EXPLANATION:
- sheet_name=None reads all sheets
- Returns a dictionary
- Keys = sheet names
- Values = DataFrames

DICTIONARY STRUCTURE:
{
    'Sales': DataFrame with sales data,
    'Profit': DataFrame with profit data,
    'Inventory': DataFrame with inventory data
}

HOW TO USE:
1. Read all sheets into dictionary
2. Access specific sheet: sheets['Sales']
3. Or loop through all sheets

SYNTAX:
all_sheets = pd.read_excel('file.xlsx', sheet_name=None)

# Access specific sheet:
sales_df = all_sheets['Sales']

# Loop through all:
for sheet_name, df in all_sheets.items():
    print(f"Processing {sheet_name}")
    print(df)
"""

all_sheets = pd.read_excel('/tmp/multi_sheet.xlsx', sheet_name=None)
print(f"\n✓ Read {len(all_sheets)} sheets")
print(f"  Sheet names: {list(all_sheets.keys())}")

for name, df in all_sheets.items():
    print(f"\n{name} sheet:")
    print(df)

# ============================================================================
# PART 4: APPENDING TO EXISTING EXCEL
# ============================================================================
print("\n" + "=" * 80)
print("PART 4: ADDING SHEETS TO EXISTING EXCEL FILE")
print("=" * 80)

"""
THE PROBLEM:
- You already have an Excel file
- Want to add a new sheet WITHOUT deleting existing ones
- Default behavior would overwrite the file

THE SOLUTION:
- Use mode='a' (append mode)
- Opens existing file
- Adds new sheet
- Keeps all existing sheets

REAL-WORLD USE CASE:
- Monthly report exists with Jan, Feb data
- Now you want to add Mar data
- Don't want to recreate Jan and Feb sheets
"""

# Create initial file
print("\nStep 1: Create initial Excel file")
df_initial = pd.DataFrame({
    'name': ['Alice', 'Bob'],
    'age': [25, 30]
})
df_initial.to_excel('/tmp/append_demo.xlsx', sheet_name='Sheet1', index=False)
print("Created file with Sheet1")

# Append new sheet
print("\nStep 2: Append new sheet to existing file")

"""
EXPLANATION:
mode='a' means append mode
- Opens existing file
- Adds new sheet
- Doesn't delete old sheets

mode='w' means write mode (default)
- Deletes entire file
- Creates new file
- Would lose Sheet1!

SYNTAX:
with pd.ExcelWriter('file.xlsx', engine='openpyxl', mode='a') as writer:
    new_df.to_excel(writer, sheet_name='Sheet2', index=False)
"""

df_new = pd.DataFrame({
    'city': ['New York', 'London'],
    'country': ['USA', 'UK']
})

with pd.ExcelWriter('/tmp/append_demo.xlsx', engine='openpyxl', mode='a') as writer:
    df_new.to_excel(writer, sheet_name='Sheet2', index=False)

print("✓ Added Sheet2 to existing file")

# Verify
all_sheets = pd.read_excel('/tmp/append_demo.xlsx', sheet_name=None)
print(f"\n✓ File now has: {list(all_sheets.keys())}")
print("  - Sheet1 still exists (not deleted)")
print("  - Sheet2 was added")

# ============================================================================
# PART 5: FORMATTING EXCEL - MAKING IT LOOK PROFESSIONAL
# ============================================================================
print("\n" + "=" * 80)
print("PART 5: FORMATTING EXCEL OUTPUT")
print("=" * 80)

"""
WHY FORMAT EXCEL FILES?
=======================
- Default Excel exports look plain
- Professional reports need formatting
- Colors, fonts, borders make data easier to read
- Impress stakeholders and clients

WHAT YOU CAN FORMAT:
- Font (bold, size, color)
- Cell colors (background)
- Borders
- Alignment
- Number formats
- Column widths

REQUIRES:
- openpyxl library
- Access to workbook and worksheet objects
"""

df_format = pd.DataFrame({
    'Product': ['Laptop', 'Mouse', 'Keyboard'],
    'Price': [1200.50, 25.99, 75.00],
    'Quantity': [10, 100, 50]
})

print("\nData to format:")
print(df_format)

print("\n" + "-" * 80)
print("EXAMPLE: Create Professional Looking Excel")
print("-" * 80)

"""
STEP-BY-STEP EXPLANATION:

1. WRITE DATA:
   - Write DataFrame to Excel
   - Use startrow=1 to leave space for title

2. ADD TITLE:
   - worksheet['A1'] = 'Sales Report'
   - Adds title in cell A1

3. FORMAT TITLE:
   - Make it bold
   - Increase font size
   - Makes it stand out

4. FORMAT HEADERS:
   - Bold font
   - Gray background color
   - Professional appearance

SYNTAX:
from openpyxl.styles import Font, PatternFill

# Bold font
cell.font = Font(bold=True, size=14)

# Background color
cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
"""

print("\nCreating formatted Excel file...")

from openpyxl.styles import Font, PatternFill

with pd.ExcelWriter('/tmp/formatted.xlsx', engine='openpyxl') as writer:
    # Write data starting from row 2 (leave row 1 for title)
    df_format.to_excel(writer, sheet_name='Products', index=False, startrow=1)
    
    # Get workbook and worksheet objects
    workbook = writer.book
    worksheet = writer.sheets['Products']
    
    # Add title in cell A1
    worksheet['A1'] = 'Sales Report'
    
    # Format title: bold and larger
    worksheet['A1'].font = Font(bold=True, size=14)
    
    # Format header row (row 2): bold with gray background
    for cell in worksheet[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

print("Formatted Excel created!")
print("\nFormatting applied:")
print("  - Title: 'Sales Report' (bold, size 14)")
print("  - Headers: Bold with gray background")
print("  - Professional appearance")

# ============================================================================
# PART 6: READING SPECIFIC RANGES - LIKE SELECTING IN EXCEL
# ============================================================================
print("\n" + "=" * 80)
print("PART 6: READING SPECIFIC RANGES")
print("=" * 80)

"""
THE CONCEPT:
- Sometimes Excel files are huge
- You only need part of the data
- Like selecting cells in Excel: A1:C10
- Saves time and memory


"""

# Create sample file
df_large = pd.DataFrame({
    'A': range(1, 101),
    'B': range(101, 201),
    'C': range(201, 301),
    'D': range(301, 401),
    'E': range(401, 501)
})
df_large.to_excel('/tmp/large.xlsx', index=False)

print(f"\nCreated file with {len(df_large)} rows and {len(df_large.columns)} columns")

# Example 1: Read first N rows
print("\n" + "-" * 80)
print("EXAMPLE 1: Read Only First 10 Rows")
print("-" * 80)

"""
EXPLANATION:
- nrows parameter limits rows
- Like "Top 10" in Excel
- Doesn't include header in count

USE CASE:
- Preview a 1 million row file
- Test your code on small sample
- Quick data check

nrows=10 means:
- Row 1: Headers
- Rows 2-11: Data (10 rows)

SYNTAX:
df = pd.read_excel('file.xlsx', nrows=10)
"""

df = pd.read_excel('/tmp/large.xlsx', nrows=10)
print(f"✓ Read only {len(df)} rows instead of 100")
print(df.head())

# Example 2: Read specific columns
print("\n" + "-" * 80)
print("EXAMPLE 2: Read Only Columns A and C")
print("-" * 80)

"""
EXPLANATION:
- usecols parameter selects columns
- Two methods: Excel letters or column names
- Like hiding columns in Excel

METHOD 1 - Excel Column Letters:
usecols='A,C'     → Columns A and C
usecols='A:C'     → Columns A through C
usecols='B:D,F'   → Columns B, C, D, and F

METHOD 2 - Column Names:
usecols=['Name', 'Age']

WHY USE THIS:
- File has 50 columns, you need 3
- Faster reading
- Less memory usage
"""

df = pd.read_excel('/tmp/large.xlsx', usecols='A,C')
print("✓ Read only columns A and C")
print(df.head())

# Example 3: Combine both
print("\n" + "-" * 80)
print("EXAMPLE 3: Read Specific Range (First 10 Rows of Columns A-C)")
print("-" * 80)


